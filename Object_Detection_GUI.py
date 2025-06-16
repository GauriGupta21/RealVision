import tkinter as tk
from tkinter import Label, Button, filedialog
import cv2
from PIL import Image, ImageTk
from ultralytics import YOLO
from datetime import datetime
import openpyxl
import uuid
from plyer import notification

triggered_objects = set()  # A set to track objects that have triggered notifications


# Load YOLO model
model = YOLO('yolov8n.pt')  # Use the YOLOv8 nano model for object detection

# Create an Excel workbook for detected objects
def create_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Detected Objects"
    sheet.append(["Timestamp", "Object", "Unique ID"])
    workbook.save("detected_objects.xlsx")

# Create a new workbook for counting object occurrences
def create_count_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Object Count"
    sheet.append(["Object", "Count"])
    workbook.save("object_count.xlsx")

# Update the count of objects in the count workbook
def update_object_count(object_name):
    workbook = openpyxl.load_workbook("object_count.xlsx")
    sheet = workbook.active
    found = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        if row[0].value == object_name:
            row[1].value += 1  
            found = True
            break
    if not found:
        sheet.append([object_name, 1])
    workbook.save("object_count.xlsx")

# Add detection details to the workbook
def log_detection(object_name):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    unique_id = str(uuid.uuid4())[:8]  
    workbook = openpyxl.load_workbook("detected_objects.xlsx")
    sheet = workbook.active
    sheet.append([timestamp, object_name, unique_id])
    workbook.save("detected_objects.xlsx")

# Function to process each frame
def process_frame(frame):
    results = model(frame)
    detections = results[0].boxes  
    detected_objects = []

    for detection in detections:
        x1, y1, x2, y2 = map(int, detection.xyxy[0])
        cls = int(detection.cls)
        confidence = detection.conf[0]

        label = f"{model.names[cls]} {confidence:.2f}"
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
        detected_objects.append(model.names[cls])
        log_detection(model.names[cls])

    # Update the object count workbook with detected objects
    for object_name in set(detected_objects):  # Avoid counting duplicates in a single frame
        update_object_count(object_name)

    return frame

def send_alert(object_name):
    # Customize this list with the objects you want alerts for
    alert_objects = ["person", "car", "bicycle"]

    # Check if the object is in the alert list and if it hasn't triggered an alert already
    if object_name.lower() in alert_objects and object_name not in triggered_objects:
        notification.notify(
            title="Object Detected",
            message=f"{object_name} detected!",
            timeout=5  # Notification stays for 5 seconds
        )
        triggered_objects.add(object_name)  # Add object to the triggered set


def process_frame(frame):
    results = model(frame)
    detections = results[0].boxes  
    detected_objects = []

    # Define custom colors for each object
    object_colors = {
        "person": (0, 0, 255),   # Red
        "cell phone": (255, 0, 0),      # Blue
        "bicycle": (0, 255, 0),  # Green
        # Add more objects and their corresponding colors as needed
    }

    for detection in detections:
        x1, y1, x2, y2 = map(int, detection.xyxy[0])
        cls = int(detection.cls)
        confidence = detection.conf[0]

        label = f"{model.names[cls]} {confidence:.2f}"
        object_name = model.names[cls]

        # Choose color based on object name, or use a default color
        color = object_colors.get(object_name, (0, 255, 255))  # Default to Yellow

        # Draw bounding box with the chosen color
        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
        cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

        detected_objects.append(object_name)
        
        # Log detection details
        log_detection(object_name)
        
        # Send an alert only if the object hasn't been triggered yet
        send_alert(object_name)

    # Update the object count workbook with detected objects
    for object_name in set(detected_objects):  # Avoid counting duplicates in a single frame
        update_object_count(object_name)

    return frame




# Creating the GUI of project using TKinter

class ObjectDetectionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Object Detection")

        self.video_source = None
        self.cap = None
        self.running = False

        self.label = Label(root, text="Object Detection", font=("Arial", 16))
        self.label.pack()

        self.canvas = tk.Canvas(root, width=800, height=600)
        self.canvas.pack()

        self.start_button = Button(root, text="Start Camera", command=self.start_camera)
        self.start_button.pack(side="left", padx=10, pady=10)

        self.select_video_button = Button(root, text="Select Video", command=self.select_video)
        self.select_video_button.pack(side="left", padx=10, pady=10)

        self.save_frame_button = Button(root, text="Save Frame", command=self.save_frame)
        self.save_frame_button.pack(side="left", padx=10, pady=10)

        self.stop_button = Button(root, text="Stop", command=self.stop_detection)
        self.stop_button.pack(side="left", padx=10, pady=10)

    def start_camera(self):
        self.video_source = 0  # Use default camera
        self.cap = cv2.VideoCapture(self.video_source)
        self.running = True
        global triggered_objects
        triggered_objects = set()
        self.update_frame()


    def select_video(self):
        self.video_source = filedialog.askopenfilename(
            title="Select Video File",
            filetypes=(("MP4 files", "*.mp4"), ("All files", "*.*")),
        )
        if self.video_source:
            self.cap = cv2.VideoCapture(self.video_source)
            self.running = True
            global triggered_objects
            triggered_objects = set()
            self.update_frame()

    def save_frame(self):
        if self.cap and self.running:
            ret, frame = self.cap.read()
            if ret:
                frame = process_frame(frame)
                filename = f"frame_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                cv2.imwrite(filename, frame)
                print(f"Frame saved as {filename}")

    def stop_detection(self):
        self.running = False
        if self.cap:
            self.cap.release()
        self.canvas.delete("all")

    def update_frame(self):
        if not self.running:
            return

        ret, frame = self.cap.read()
        if ret:
            frame = process_frame(frame)

            # Calculate the aspect ratio of the frame and scale to fit the canvas
            height, width, _ = frame.shape
            canvas_width = 800
            canvas_height = 600
            aspect_ratio = width / height

            if aspect_ratio > 1:  
                new_width = canvas_width
                new_height = int(canvas_width / aspect_ratio)
            else:  
                new_height = canvas_height
                new_width = int(canvas_height * aspect_ratio)
            resized_frame = cv2.resize(frame, (new_width, new_height))
            pad_top = (canvas_height - new_height) // 2
            pad_left = (canvas_width - new_width) // 2
            padded_frame = cv2.copyMakeBorder(
                resized_frame, pad_top, pad_top, pad_left, pad_left, cv2.BORDER_CONSTANT, value=(0, 0, 0)
            )

            padded_frame = cv2.cvtColor(padded_frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(padded_frame)
            imgtk = ImageTk.PhotoImage(image=img)

            self.canvas.create_image(0, 0, anchor="nw", image=imgtk)
            self.root.imgtk = imgtk

        self.root.after(10, self.update_frame)

create_workbook()
create_count_workbook()

root = tk.Tk()
app = ObjectDetectionApp(root)
root.mainloop()
